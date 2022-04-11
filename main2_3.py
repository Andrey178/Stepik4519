import openpyxl
from main2_1 import get_xls_file

filename, file_url = 'trekking1.xlsx', 'https://stepik.org/media/attachments/lesson/245290/trekking1.xlsx'


def load_openpyxl(xls_filename):  # Open xlsx file with openpyxl library and count
    wb = openpyxl.load_workbook(xls_filename, data_only=True, keep_links=False)
    sheet_names = wb.sheetnames
    # print(sheet_names)
    sh = wb[sheet_names[0]]
    return sh


def sort_caloric(xls_filename):  # Sort data in table and print it
    table = load_openpyxl(xls_filename)
    menu = []
    for row in table.iter_rows(min_row=2, max_col=2, values_only=True):
        menu.append(row)
    menu.sort(key=(lambda food: food[0]))
    menu.sort(key=(lambda food: -food[1]), reverse=True)
    for elem in menu:
        print(elem[0])


def task(name):
    get_xls_file(filename, file_url)
    #   load_xlrd(filename)
    sort_caloric(filename)


if __name__ == '__main__':
    task('main')
