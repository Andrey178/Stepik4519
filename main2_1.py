import openpyxl
import wget
import xlrd

filename, file_url = 'salaries.xlsx', 'https://stepik.org/media/attachments/lesson/245267/salaries.xlsx'


def get_xls_file(xls_filename, xls_file_url):  # Download xlsx file from source
    try:
        with open(xls_filename):
            pass
    except FileNotFoundError:
        xls_filename = wget.download(xls_file_url, xls_filename)
    return xls_filename


def load_xlrd(xls_filename):  # Open xlsx file with xlrd library and count
    wb = xlrd.open_workbook(xls_filename)
    sheet_names = wb.sheet_names()
    # print(sheet_names)
    sh = wb.sheet_by_name(sheet_names[0])
    n_min = sh.row_values(6)[2]
    for row_num in range(7, 27):
        temp = sh.row_values(row_num)
        n_min = min(n_min, temp[2])
    print(n_min)


def load_openpyxl(xls_filename):  # Open xlsx file with openpyxl library and count
    wb = openpyxl.load_workbook(xls_filename, data_only=True, keep_links=False)
    sheet_names = wb.sheetnames
    # print(sheet_names)
    sh = wb[sheet_names[0]]

    # find_min(sh)
    find_mean_salary(sh)
    find_median_salary(sh)


def find_median_salary(sheet):
    region_salaries = list(sheet.values)[1:]
    rich_region, region_med_salary = '', 0
    for row in region_salaries:
        med_salary = sorted(row[1:])[3]
        if region_med_salary < med_salary:
            rich_region, region_med_salary = row[0], med_salary

    print(f"{rich_region} : {region_med_salary}")


def find_min(sheet):
    print(type(sheet))
    n_min = sheet.cell(row=7, column=3).value
    print(f"{str(n_min)} => {str(sheet.cell(row=7, column=3).coordinate)}")
    for row_num in sheet['C7':'C27']:
        temp = row_num[0].value
        n_min = min(n_min, temp)
    print(n_min)


def find_mean_salary(sheet):
    prof_salary = {}
    for col in range(2, 9):
        prof_name = sheet.cell(column=col, row=1).value
        prof_salary[prof_name] = sheet.cell(column=col, row=2).value
        for row in range(3, 10):
            prof_salary[prof_name] = prof_salary[prof_name] + sheet.cell(column=col, row=row).value
    best_prof, best_salary = '', 0
    for elem in prof_salary.items():
        if elem[1] / 8 > best_salary:
            best_prof = elem[0]
            best_salary = elem[1] / 8
    print(f"{best_prof} : {best_salary}")


def task(name):
    get_xls_file(filename, file_url)
    #   load_xlrd(filename)
    load_openpyxl(filename)


if __name__ == '__main__':
    task('main')
