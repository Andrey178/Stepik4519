import openpyxl as openpyxl
import requests
import wget
import xlrd


def task(name):
    # Download xlsx file from source
    filename, file_url = 'tab.xlsx', 'https://stepik.org/media/attachments/lesson/245266/tab.xlsx'
    # filename = wget.download(, 'tab.xlsx')
    print(filename)
    file_bin = requests.get(file_url)

    # Open xlsx file with xlrd library and count
    def load_xlrd(xls_filename):
        wb = xlrd.open_workbook(file_contents=xls_filename.content)
        sheet_names = wb.sheet_names()
        print(sheet_names)
        sh = wb.sheet_by_name(sheet_names[0])
        n_min = sh.row_values(6)[2]
        for row_num in range(7, 27):
            temp = sh.row_values(row_num)
            n_min = min(n_min, temp[2])
        print(n_min)

    # Open xlsx file with openpyxl library and count
    def load_openpyxl(xls_filename):
        try:
            with open(xls_filename):
                pass
        except FileNotFoundError:
            xls_filename = wget.download(file_url, 'tab.xlsx')

        wb = openpyxl.load_workbook(xls_filename)
        sheet_names = wb.sheetnames
        print(sheet_names)
        sh = wb[sheet_names[0]]
        n_min = sh.cell(row=7, column=3).value
        print(f"{str(n_min)} => {str(sh.cell(row=7, column=3).coordinate)}")
        for row_num in sh['C7':'C27']:
            temp = row_num[0].value
            n_min = min(n_min, temp)
        print(n_min)

    load_xlrd(file_bin)
    load_openpyxl(filename)


if __name__ == '__main__':
    task('main')
