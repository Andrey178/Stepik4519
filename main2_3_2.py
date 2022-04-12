import openpyxl
from main2_1 import get_xls_file

filename, pagename, file_url = 'trekking2.xlsx', ['Справочник', 'Раскладка'],\
                               'https://stepik.org/media/attachments/lesson/245290/trekking2.xlsx'


def load_openpyxl_file(xls_filename, page=0):  # Open xlsx file with openpyxl library and count
    wb = openpyxl.load_workbook(xls_filename, data_only=True, keep_links=False)
    return wb


def count_product_data(xls_filename, pagename):  # Sort data in table and print it
    table = load_openpyxl_file(xls_filename)
    # print(table.sheetnames)
    sh0 = table[pagename[0] if pagename[0] in table.sheetnames else table.sheetnames[0]]
    sh1 = table[pagename[1] if pagename[1] in table.sheetnames else table.sheetnames[1]]
    # day_menu, menu = [], list(sh0.iter_rows(min_row=2, max_col=5, values_only=True))
    day_menu, menu = [], dict((elem[0], elem[1:]) for elem in sh0.iter_rows(min_row=2, max_col=5, values_only=True))
    for row in sh1.iter_rows(min_row=2, max_col=2, values_only=True):  # Take line from day menu
        menu_data = list(row)
        menu_data.extend(menu[menu_data[0]])
        day_menu.append(menu_data)

    day_menu_calories = [0]*6
    for elem in day_menu:
        for index in range(2, len(elem)):
            product_item = elem[index] or 0
            day_menu_calories[index] += product_item*(elem[1]/100)
    # print(*list(map(lambda elem1: int(elem1), day_menu_calories[2:])))
    print(*(int(elem) for elem in day_menu_calories[2:]))


def task(name):
    get_xls_file(filename, file_url)
    count_product_data(filename, pagename)


if __name__ == '__main__':
    task('main')
